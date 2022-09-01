import base64
import os
import random
from datetime import date, datetime
import pytz
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from termcolor import colored
from selenium.webdriver.common.action_chains import ActionChains
from configparser import ConfigParser
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
import re

# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#                                                   USER INPUT
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
CONFIG_FILE_PATH = "G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini"
# CUSTOMER_ID = [150, 200, 1000, 1100, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900, 2000, 2100, 2200, 2400, 2500, 2600,
#                2700, 2800, 2900, 3000, 3100, 3200, 3300, 3400, 3500, 3600, 3700, 3800, 3900, 4000, 4100, 4200, 4300,
#                4400, 4500, 4600, 4700, 4800, 4900, 5000, 5100, 5200, 5300, 5400, 5600, 6000, 6500, 5900]
CUSTOMER_ID = [150, 200, 1300, 3300, 3600]


# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def ChartList():
    """ CONFIG PARSER """
    file = '../config.ini'
    config = ConfigParser()
    config.read(CONFIG_FILE_PATH)

    """ CHROME DRIVER SETUP """
    options = webdriver.ChromeOptions()
    options.add_argument(config['path']['chrome_profile'])  # Path to your chrome profile
    # options.add_argument("--headless")
    # options.add_argument('--disable-gpu')
    # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--start-maximized")
    driver = webdriver.Chrome(executable_path=(config['path']['chrome_driver']), options=options)

    def date_time():
        today = date.today()
        tz_In = pytz.timezone('Asia/Kolkata')
        datetime_In = datetime.now(tz_In)
        time_now = datetime_In.strftime("[%I-%M-%S %p]")
        now = str(today) + time_now
        print(now)
        return now

    """ REPORT FILE CREATION """
    wb = Workbook()
    wb.active.title = "ChartsList "
    sh1 = wb.active
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    sh1['A1'].font = header_font
    sh1['A1'].fill = header_cell_color
    sh1['B1'].font = header_font
    sh1['B1'].fill = header_cell_color
    sh1['C1'].font = header_font
    sh1['C1'].fill = header_cell_color
    sh1['D1'].font = header_font
    sh1['D1'].fill = header_cell_color
    sh1['E1'].font = header_font
    sh1['E1'].fill = header_cell_color
    sh1['F1'].font = header_font
    sh1['F1'].fill = header_cell_color
    sh1['G1'].font = header_font
    sh1['G1'].fill = header_cell_color

    sh1['A1'] = 'Customer Name'
    sh1['B1'] = 'Region Name'
    sh1['C1'] = 'Chart List Name'
    sh1['D1'] = 'Time Taken(default date filter)'
    sh1['E1'] = 'Time Taken(date: 01/01/2015 to current date)'
    sh1['F1'] = 'Comments'
    sh1['G1'] = 'URL'

    sh1.name = "Arial"

    """ FOLDER CREATION """
    dateandtime = date_time()
    path = os.path.join((config['path']['parent_dir']), dateandtime + "_ChartsList")
    os.mkdir(path)
    path1 = os.path.join(path, "Charts List")
    os.mkdir(path1)
    path2 = os.path.join(path, "Screenshots")
    os.mkdir(path2)

    # LOGGING AREA
    LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
    logging.basicConfig(filename=path1 + "\\" + "Info.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    wb.save(path1 + '\\' + "Report.xlsx")

    def Login():
        try:
            driver.get(config['prod']['logout_url'])
            driver.get(config['prod']['login_url'])
            driver.maximize_window()
            driver.find_element_by_id("edit-name").send_keys(config['credentials']['username'])
            driver.find_element_by_id("edit-pass").send_keys(config['credentials']['password'])
            driver.find_element_by_id("edit-submit").click()
            time.sleep(2)
            WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            driver.find_element_by_id("reason_textbox").send_keys(config['credentials']['reason'])
            time.sleep(0.5)
            driver.find_element_by_id("edit-submit").click()
            ajax_preloader_wait()
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_filterMeasureList'])))
            logger.info("Logged in to Cozeva!")

        except Exception as e:
            print(e)
            logger.critical("Exception occurred in Login function!")
            raise

    def ajax_preloader_wait():
        try:
            time.sleep(1)
            WebDriverWait(driver, 120).until(EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
            WebDriverWait(driver, 120).until(EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))
            time.sleep(1)

        except Exception as e:
            return e

    def open_registry_page(customer_id):
        customer_list_url = []
        sm_customer_id = str(customer_id)
        sm_customer_id = sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        customer_list_url.append(encoded_string)
        for idx, val in enumerate(customer_list_url):
            # driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))
            driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))
        ajax_preloader_wait()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_contextName'])))
        customer_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
        return customer_name

    def open_region_registry_page():
        global region_name
        ajax_preloader_wait()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//div[@id='list_1']/ul/li")))
        time.sleep(5)
        available_redirect_links = driver.find_elements(by=By.XPATH, value="//div[@id='list_1']/ul/li")

        for current_redirect_link_counter in range(len(available_redirect_links)):
            current_redirect_link = available_redirect_links[current_redirect_link_counter].get_attribute("redirect_link")
            print(current_redirect_link)
            driver.get("https://www.cozeva.com" + current_redirect_link)
            ajax_preloader_wait()
            time.sleep(5)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_contextName'])))
            region_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
            open_supplemental_data_list()
            time.sleep(5)
            available_redirect_links = driver.find_elements(by=By.XPATH, value="//div[@id='list_1']/ul/li")

    def open_supplemental_data_list():
        chart_list_redirect_links = driver.find_elements(by=By.XPATH, value="//li[contains(@class, 'chart_chase_list_type')]/a")

        for chart_list_redirect_link_counter in range(len(chart_list_redirect_links)):
            chart_list_redirect_link = chart_list_redirect_links[chart_list_redirect_link_counter].get_attribute("href")
            print(chart_list_redirect_link)

            try:
                driver.get(chart_list_redirect_link)
                start_time = None
                start_time = time.perf_counter()
                ajax_preloader_wait()
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[@class='datatable_filter_dropdown sidenav-trigger']")))
                time_taken_default = round((time.perf_counter() - start_time - 2), 2)
                chart_list_name = driver.find_element(by=By.XPATH, value="//div[@class='ch table_header']").text
                driver.find_element(by=By.XPATH, value="//a[@class='datatable_filter_dropdown sidenav-trigger']").click()
                time.sleep(1.5)
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//input[@title='Created']")))
                driver.find_element(by=By.XPATH, value="//input[@title='Created']").clear()
                driver.find_element(by=By.XPATH, value="//input[@title='Created']").send_keys("01/01/2015")
                time.sleep(1)
                driver.find_element(by=By.XPATH, value="//div[@class='clearfix']/a[contains(text(),'Apply')]").click()
                time.sleep(1)
                start_time = None
                start_time = time.perf_counter()
                ajax_preloader_wait()
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[@class='datatable_filter_dropdown sidenav-trigger']")))
                time_taken_date_filter_set = round((time.perf_counter() - start_time - 2), 2)
                sh1.append((customer_name, region_name, chart_list_name, time_taken_default, time_taken_date_filter_set, '', ''))

            except Exception as e:
                print(e)
                sh1.append((customer_name, region_name, chart_list_name, '0', '0', 'Exception Occurred!', chart_list_redirect_link))

            rows = sh1.max_row
            cols = sh1.max_column
            for i in range(2, rows + 1):
                for j in range(4, 6):
                    try:
                        if sh1.cell(i, j).value == "Exception Occurred!":
                            sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
                        elif float(sh1.cell(i, j).value) > 10 and float(sh1.cell(i, j).value) < 30:
                            sh1.cell(i, j).fill = PatternFill('solid', fgColor='FDD3C9')
                        elif float(sh1.cell(i, j).value) >= 30:
                            sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
                    except:
                        continue

            wb.save(path1 + '\\' + "Report.xlsx")

            chart_list_redirect_links = driver.find_elements(by=By.XPATH, value="//li[contains(@class, 'chart_chase_list_type')]/a")

    Login()
    for cust_id_counter in CUSTOMER_ID:
        customer_name = open_registry_page(cust_id_counter)
        open_region_registry_page()

    driver.get(config['prod']['logout_url'])
    driver.quit()


ChartList()