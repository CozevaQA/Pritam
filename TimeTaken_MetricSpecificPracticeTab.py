import base64
import os
import random
from datetime import date, datetime
import pytz
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from termcolor import colored
from selenium.webdriver.common.action_chains import ActionChains
from configparser import ConfigParser
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
import re

# +++++++++++++++++++++++++++++++++++++ USER INPUT ++++++++++++++++++++++++++++++++++++++++++++++++++++++++
CONFIG_FILE_PATH = "G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini"
Customer_Id = [2100, 200]


# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def BlankList_Support_Script(CUSTOMER_ID, QUARTER, COUNTER):
    # CONFIG PARSER
    file = '../config.ini'
    config = ConfigParser()
    config.read(CONFIG_FILE_PATH)

    """ CHROME DRIVER SETUP """
    options = webdriver.ChromeOptions()
    options.add_argument(config['path']['chrome_profile'])  # Path to your chrome profile
    # options.add_argument("--headless")
    # options.add_argument('--disable-gpu')
    # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--start-maximized")
    driver = webdriver.Chrome(executable_path=(config['path']['chrome_driver']), options=options)

    def date_time():
        today = date.today()
        tz_In = pytz.timezone('Asia/Kolkata')
        datetime_In = datetime.now(tz_In)
        time_now = datetime_In.strftime("[%I-%M-%S %p]")
        now = str(today) + time_now
        print(now)
        return now

    wb = Workbook()
    wb.active.title = "CUSTOMER_ID " + str(CUSTOMER_ID)
    sh1 = wb.active
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    sh1['A1'].font = header_font
    sh1['A1'].fill = header_cell_color
    sh1['B1'].font = header_font
    sh1['B1'].fill = header_cell_color
    sh1['C1'].font = header_font
    sh1['C1'].fill = header_cell_color
    sh1['D1'].font = header_font
    sh1['D1'].fill = header_cell_color
    sh1['E1'].font = header_font
    sh1['E1'].fill = header_cell_color
    sh1['F1'].font = header_font
    sh1['F1'].fill = header_cell_color
    sh1['G1'].font = header_font
    sh1['G1'].fill = header_cell_color
    sh1['H1'].font = header_font
    sh1['H1'].fill = header_cell_color
    sh1['I1'].font = header_font
    sh1['I1'].fill = header_cell_color
    sh1['J1'].font = header_font
    sh1['J1'].fill = header_cell_color
    sh1['A1'] = 'Quarter & LOB'
    sh1['B1'] = 'Measure Name'
    sh1['C1'] = 'Numerator'
    sh1['D1'] = 'Denominator'
    sh1['E1'] = 'Tab Name'
    sh1['F1'] = 'Count'
    sh1['G1'] = 'Time taken'
    sh1['H1'] = 'Status'
    sh1['I1'] = 'Comments'
    sh1['J1'] = 'URL'
    sh1.name = "Arial"

    # **** FOLDER CREATION ****
    dateandtime = date_time()
    path = os.path.join((config['path']['parent_dir']), dateandtime + "_Support level_Customer " + str(CUSTOMER_ID))
    os.mkdir(path)
    path1 = os.path.join(path, "Cust Id " + str(CUSTOMER_ID) + "_" + "SupportList")
    os.mkdir(path1)
    path2 = os.path.join(path, "Screenshots")
    os.mkdir(path2)

    # **** LOGGING AREA ****
    LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
    logging.basicConfig(filename=path1 + "\\" + "Info.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    wb.save(path1 + '\\' + "Report.xlsx")

    def Login():
        try:
            driver.get(config['prod']['logout_url'])
            driver.get(config['prod']['login_url'])
            driver.maximize_window()
            driver.find_element_by_id("edit-name").send_keys(config['credentials']['username'])
            driver.find_element_by_id("edit-pass").send_keys(config['credentials']['password'])
            driver.find_element_by_id("edit-submit").click()
            time.sleep(2)
            WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            driver.find_element_by_id("reason_textbox").send_keys(config['credentials']['reason'])
            time.sleep(0.5)
            driver.find_element_by_id("edit-submit").click()
            ajax_preloader_wait()
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_filterMeasureList'])))
            logger.info("Logged in to Cozeva!")

        except Exception as e:
            print(e)
            logger.critical("Exception occurred in Login function!")
            raise

    def support_blankList():
        global providers_list_url
        open_registry_page(CUSTOMER_ID)
        ajax_preloader_wait()
        customer_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
        print("**** " + customer_name + " ****")
        driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
        time.sleep(1)
        quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")
        lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
        driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()

        for quarter in range(QUARTER):
            # quarter = quarter + 1
            for lob in range(len(lobs)):
                # for lob in range(3):
                # lob = lob + 4
                time.sleep(0.5)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@id='qt-filter-label']")))
                driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
                time.sleep(0.25)
                quarter_name = quarters[quarter].text
                print(colored(quarter_name, 'blue'))
                quarters[quarter].click()
                time.sleep(0.25)
                lobs[lob].click()
                lob_name = lobs[lob].text
                lob_name = re.sub("[\/:*?<>|]", "", lob_name)
                print(colored(lob_name, 'magenta'))
                driver.find_element_by_xpath("//a[@id='reg-filter-apply']").click()
                ajax_preloader_wait()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))
                driver.find_element_by_xpath("//a[@data-target='qt-reg-nav-filters']").click()
                time.sleep(0.25)
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//label[@class='col s12']")))
                driver.find_element_by_xpath("//label[@class='col s12']").click()
                time.sleep(0.25)
                driver.find_element_by_xpath("//button[@id='qt-apply-search']").click()
                ajax_preloader_wait()

                """
                **** SUPPORT MEASURE REGISTRY NAVIGATION ****
                """
                measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                measure_counter = 0
                score = 0

                while measure_counter < len(measures_all) and score < len(scores):
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, "//a[@id='reg-faq-trigger']")))
                    time.sleep(0.5)

                    driver.execute_script("arguments[0].scrollIntoView();", measures_all[measure_counter])
                    measure_name = (measures_all[measure_counter]).text
                    measure_name = re.sub("[\/:*?<>|]", "", measure_name)
                    print("Measure name: ", measure_name)
                    numdeno = scores[score].text
                    numdeno = numdeno.lstrip("(")
                    numdeno = numdeno.rstrip(")")
                    numdeno = numdeno.split("/")
                    Numerator = numdeno[0].replace(',', '')
                    Denominator = numdeno[1].replace(',', '')
                    print("Numerator=", Numerator)
                    print("Denominator=", Denominator)
                    last_url = driver.current_url
                    measures_all[measure_counter].click()

                    try:
                        time.sleep(0.5)
                        providers_list_url = driver.current_url
                        """ PROVIDERS LIST """
                        # start_time = None
                        # end_time = None
                        # start_time = time.time()
                        ajax_preloader_wait()
                        # end_time = time.time()
                        # rows = len(driver.find_elements_by_xpath("//table[@id='metric-support-prov-ls']/tbody/tr[@role='row']"))
                        # print(measure_name+"~"+str(rows)+"~"+str(end - start))
                        # print("Manual intervention needed")

                        # time_taken_providers = round((end_time - start_time), 2)
                        # provider_count = len(driver.find_elements_by_xpath("//tbody/tr[@role='row']"))
                        # prov_datatable_info = driver.find_element_by_xpath("//div[@id='metric-support-prov-ls_info']").text
                        # if provider_count == 0:
                        #     if float(Denominator) != 0:
                        #         sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator, 'Providers', prov_datatable_info, time_taken_providers, 'FAILED', 'List is blank though non-zero measure score!', providers_list_url ))
                        #         driver.save_screenshot(path2 + "\\" + quarter_name + " "+lob_name + "_"+ measure_name +"_Provider" + ".png")

                        #     elif float(Denominator) == 0:
                        #         if float(Numerator) == 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,'Providers', provider_count, time_taken_providers, 'NA', 'List is blank' ))
                        #             driver.save_screenshot(path2 + "\\" + quarter_name + " "+lob_name + "_"+ measure_name+"_Provider" + ".png")
                        #         elif float(Numerator) != 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Providers', provider_count, time_taken_providers, 'FAILED', 'List is blank though non-zero measure score!', providers_list_url ))
                        #             driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+"_Provider" + ".png")

                        # elif provider_count != 0:
                        #     if float(Denominator) != 0:
                        #         sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Providers', prov_datatable_info, time_taken_providers, 'PASSED', ''))
                        #         driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+"_Provider" + ".png")
                        #     elif float(Denominator) == 0:
                        #         if float(Numerator) == 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Providers', provider_count, time_taken_providers, 'FAILED', 'List is populated though measure score is zero!', providers_list_url ))
                        #             driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+"_Provider" + ".png")
                        #         elif float(Numerator) != 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Providers', prov_datatable_info, time_taken_providers, 'PASSED', ''))
                        #             driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+"_Provider" + ".png")

                        """ PATIENTS LIST """
                        # driver.find_element_by_xpath("(//ul[@id='qt-mt-support-ls']/li)[3]").click()
                        # patients_list_url = driver.current_url
                        # start_time = None
                        # end_time = None
                        # start_time = time.time()
                        # ajax_preloader_wait()
                        # end_time = time.time()
                        # # rows = len(driver.find_elements_by_xpath("//table[@id='metric-support-prov-ls']/tbody/tr[@role='row']"))
                        # # print(measure_name+"~"+str(rows)+"~"+str(end - start))
                        # # print("Manual intervention needed")
                        # time_taken_patients = round((end_time - start_time), 2)
                        # patient_count = len(driver.find_elements_by_xpath("//tbody/tr[@role='row']"))
                        # pat_datatable_info = driver.find_element_by_xpath("//div[@id='metric-support-pat-ls_info']").text
                        # if patient_count == 0:
                        #     if float(Denominator) != 0:
                        #         sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Patients', pat_datatable_info, time_taken_patients, 'FAILED', 'List is blank though non-zero measure score!', patients_list_url ))
                        #         driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+"_Patients" + ".png")

                        #     elif float(Denominator) == 0:
                        #         if float(Numerator) == 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,'Patients', patient_count, time_taken_patients, 'NA', 'List is blank' ))
                        #             driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+ "_Patients" + ".png")
                        #         elif float(Numerator) != 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,'Patients', patient_count, time_taken_patients, 'FAILED', 'List is blank though non-zero measure score!', patients_list_url))
                        #             driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+ "_Patients" + ".png")

                        # elif patient_count != 0:
                        #     if float(Denominator) != 0:
                        #         sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Patients', pat_datatable_info, time_taken_patients, 'PASSED', ''))
                        #         driver.save_screenshot(path2 + "\\" + quarter_name + " "+lob_name + "_"+ measure_name+"_Patients" +  ".png")
                        #     elif float(Denominator) == 0:
                        #         if float(Numerator) == 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,'Patients', patient_count, time_taken_patients, 'FAILED', 'List is populated though measure score is zero!', patients_list_url ))
                        #             driver.save_screenshot(path2 + "\\" +quarter_name + " "+ lob_name + "_"+ measure_name+ "_Patients" + ".png")
                        #         elif float(Numerator) != 0:
                        #             sh1.append((quarter_name + " | " + lob_name, measure_name,Numerator, Denominator, 'Patients', pat_datatable_info, time_taken_patients, 'PASSED', ''))
                        #             driver.save_screenshot(path2 + "\\" + quarter_name + " "+lob_name + "_"+ measure_name+ "_Patients" + ".png")

                        """ ******** PRACTICES LIST ********"""
                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.XPATH, "(//ul[@id='qt-mt-support-ls']/li)[1]")))
                        driver.find_element_by_xpath("(//ul[@id='qt-mt-support-ls']/li)[1]").click()
                        practice_list_url = driver.current_url
                        start_time = None
                        start_time = time.perf_counter()
                        ajax_preloader_wait()
                        time_taken_practice = round((time.perf_counter() - start_time - 2), 2)
                        # time_taken_practice = round(time_taken_practiceToggle, 2)
                        practice_count = len(driver.find_elements_by_xpath("//tbody/tr[@role='row']"))
                        prac_datatable_info = driver.find_element_by_xpath(
                            "//div[@id='metric-support-prac-ls_info']").text

                        if practice_count == 0:
                            if float(Denominator) != 0:
                                sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,
                                            'practice', prac_datatable_info, time_taken_practice, 'FAILED',
                                            'List is blank though non-zero measure score!', practice_list_url))
                                driver.save_screenshot(
                                    path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_practice" + ".png")

                            elif float(Denominator) == 0:
                                if float(Numerator) == 0:
                                    sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,
                                                'practice', practice_count, time_taken_practice, 'NA', 'List is blank'))
                                    driver.save_screenshot(
                                        path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_practice" + ".png")
                                elif float(Numerator) != 0:
                                    sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,
                                                'practice', practice_count, time_taken_practice, 'FAILED',
                                                'List is blank though non-zero measure score!', practice_list_url))
                                    driver.save_screenshot(
                                        path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_practice" + ".png")

                        elif practice_count != 0:
                            if float(Denominator) != 0:
                                sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,
                                            'practice', prac_datatable_info, time_taken_practice, 'PASSED', ''))
                                driver.save_screenshot(
                                    path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_practice" + ".png")
                            elif float(Denominator) == 0:
                                if float(Numerator) == 0:
                                    sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,
                                                'practice', practice_count, time_taken_practice, 'FAILED',
                                                'List is populated though measure score is zero!', practice_list_url))
                                    driver.save_screenshot(
                                        path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_practice" + ".png")
                                elif float(Numerator) != 0:
                                    sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator,
                                                'practice', prac_datatable_info, time_taken_practice, 'PASSED', ''))
                                    driver.save_screenshot(
                                        path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_practice" + ".png")

                        """ **************************************************** PRACTICE LIST END ****************************************************"""

                        WebDriverWait(driver, 120).until(
                            EC.presence_of_element_located((By.XPATH, "//a[@class='breadcrumb']")))
                        driver.find_element_by_xpath("//a[@class='breadcrumb']").click()

                    except Exception as e:
                        print(e)
                        sh1.append((quarter_name + " | " + lob_name, measure_name, Numerator, Denominator, '', '', '',
                                    'FAILED', 'Exception occurred!', providers_list_url))
                        driver.save_screenshot(
                            path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + "_Exception" + ".png")

                        # WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@class='breadcrumb']")))
                        # driver.find_element_by_xpath("//a[@class='breadcrumb']").click()

                        """#DEBUG"""
                        driver.get(last_url)
                        WebDriverWait(driver, 90).until(
                            EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))

                        driver.find_element_by_xpath("//a[@data-target='qt-reg-nav-filters']").click()
                        time.sleep(0.25)
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//label[@class='col s12']")))

                        driver.find_element_by_xpath("//label[@class='col s12']").click()
                        time.sleep(0.25)

                        driver.find_element_by_xpath("//button[@id='qt-apply-search']").click()
                        WebDriverWait(driver, 90).until(
                            EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))

                        """#DEBUG"""
                    finally:
                        measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                        scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                        measure_counter += COUNTER
                        score += COUNTER
                        rows = sh1.max_row
                        cols = sh1.max_column
                        for i in range(2, rows + 1):
                            for j in range(2, cols + 1):
                                if sh1.cell(i, j).value == 'PASSED':
                                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                                elif sh1.cell(i, j).value == "FAILED":
                                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')

                        wb.save(path1 + '\\' + "Report.xlsx")

                lobs = driver.find_elements_by_xpath(
                    "//ul[@id='filter-lob']/li[@class!='hide']")  # //ul[@id='filter-lob']/li

                quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")

    def ajax_preloader_wait():
        try:
            time.sleep(1)
            WebDriverWait(driver, 300).until(
                EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
            time.sleep(1)

        except Exception as e:
            return e

    def open_registry_page(customer_id):
        customer_list_url = []
        sm_customer_id = str(customer_id)
        sm_customer_id = sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        customer_list_url.append(encoded_string)
        for idx, val in enumerate(customer_list_url):
            # driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))
            driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))

    Login()
    support_blankList()
    driver.get(config['prod']['logout_url'])
    driver.quit()


for cust_id_counter in Customer_Id:
    BlankList_Support_Script(cust_id_counter, 1, 1)
    # BlankList_Support_Script(2100, 2, 3)
    # BlankList_Support_Script(2000, 2, 2)
