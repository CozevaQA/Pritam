import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
from datetime import date, datetime
import pytz
from configparser import ConfigParser

CONFIG_FILE_PATH = "G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini"


class ConfParser:
    def __init__(self, config_file_path):
        self.config_file_path = config_file_path
        self.file = '../config.ini'
        self.config = ConfigParser()
        self.config.read(self.config_file_path)
        print("<Execution Log> Config Parser read data from:", self.config_file_path + "\n")


class FolderCreation(ConfParser):
    def date_time(self):
        self.today = date.today()
        self.tz_In = pytz.timezone('Asia/Kolkata')
        self.datetime_In = datetime.now(self.tz_In)
        self.time = self.datetime_In.strftime("[%I-%M-%S %p]")
        self.now = str(self.today) + self.time
        print("<Execution Log> ", self.now)
        return self.now

    def folder_create(self):
        global sh1, sh2, sh3, sh4, wb, path1
        wb = Workbook()
        wb.active.title = "PatientsCount"
        sh1 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh1['A1'].font = header_font
        sh1['A1'].fill = header_cell_color
        sh1['B1'].font = header_font
        sh1['B1'].fill = header_cell_color
        sh1['C1'].font = header_font
        sh1['C1'].fill = header_cell_color
        sh1['D1'].font = header_font
        sh1['D1'].fill = header_cell_color
        sh1['E1'].font = header_font
        sh1['E1'].fill = header_cell_color
        sh1['A1'] = 'Customer Name'
        sh1['B1'] = 'Provider Name'
        sh1['C1'] = 'Previous Patient Count'
        sh1['D1'] = 'Current Patient Count'
        sh1['E1'] = '% Drop'
        sh1.name = "Arial"

        sh2 = wb.create_sheet(title='DenominatorCount')
        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'DenominatorCount':
                break
        wb.active = sheet_counter
        sh2 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh2['A1'].font = header_font
        sh2['A1'].fill = header_cell_color
        sh2['B1'].font = header_font
        sh2['B1'].fill = header_cell_color
        sh2['C1'].font = header_font
        sh2['C1'].fill = header_cell_color
        sh2['D1'].font = header_font
        sh2['D1'].fill = header_cell_color
        sh2['A1'] = 'Customer Name'
        sh2['B1'] = 'Provider Name'
        sh2['C1'] = 'Previous Denominator'
        sh2['D1'] = 'Current Denominator'
        sh2.name = "Arial"

        sh3 = wb.create_sheet(title='Deleted')
        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'Deleted':
                break
        wb.active = sheet_counter
        sh3 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh3['A1'].font = header_font
        sh3['A1'].fill = header_cell_color
        sh3['B1'].font = header_font
        sh3['B1'].fill = header_cell_color
        sh3['C1'].font = header_font
        sh3['C1'].fill = header_cell_color
        sh3['A1'] = 'Customer Name'
        sh3['B1'] = 'Provider Name'
        sh3['C1'] = 'Previous Patients Count'
        sh3.name = "Arial"

        sh4 = wb.create_sheet(title='Added')
        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'Added':
                break
        wb.active = sheet_counter
        sh4 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh4['A1'].font = header_font
        sh4['A1'].fill = header_cell_color
        sh4['B1'].font = header_font
        sh4['B1'].fill = header_cell_color
        sh4['C1'].font = header_font
        sh4['C1'].fill = header_cell_color
        sh4['A1'] = 'Customer Name'
        sh4['B1'] = 'Provider Name'
        sh4['C1'] = 'Current Patients Count'
        sh4.name = "Arial"

        self.dateandtime = self.date_time()
        print("D1:", self.config_file_path)
        path = os.path.join((self.config['path']['pandas_dir']), self.dateandtime)
        os.mkdir(path)
        path1 = os.path.join(path, "Report_Provider")
        os.mkdir(path1)
        wb.save(path1 + '\\' + "Report_Provider.xlsx")
        print("<Execution Log> Folder Created")


class CompareProvidersList(FolderCreation):
    def compare_prov_prev_vs_cur(self):
        df_prev = pd.read_csv("C:\\Users\\psur\Documents\\PandasPerformance\\PandasInput\\Provider\\ProviderList_StJoseph_16Nov.csv", skiprows=1)
        # print(df_prev)

        df_cur = pd.read_csv("C:\\Users\\psur\Documents\\PandasPerformance\\PandasInput\\Provider\\ProviderList_StJoseph_17Nov.csv", skiprows=1)
        # print(df_cur)

        df = pd.merge(df_prev, df_cur, on=' Provider ID', how="outer", indicator=True, suffixes=('_prev', '_cur'))
        print(df)

        for index in range(len(df)):
            if df.loc[index, '_merge'] == "both":
                if int(df.loc[index, "Patients_prev"]) > 50:
                    for sheet_counter in range(len(wb.sheetnames)):
                        if wb.sheetnames[sheet_counter] == 'PatientsCount':
                            break
                    wb.active = sheet_counter
                    drop_percentage = round(float((int(df.loc[index, "Patients_prev"]) - int(df.loc[index, "Patients_cur"])) * 100 / int(df.loc[index, "Patients_prev"])))
                    if int(df.loc[index, "Patients_cur"]) == 0:
                        print(df.loc[index, "Name_prev"] + ": Patient count becomes zero!")
                        sh1.append((" ", df.loc[index, "Name_prev"], df.loc[index, "Patients_prev"], df.loc[index, "Patients_cur"], drop_percentage))

                    elif drop_percentage > 50 and drop_percentage < 100:
                        sh1.append((" ", df.loc[index, "Name_prev"], df.loc[index, "Patients_prev"], df.loc[index, "Patients_cur"], drop_percentage))

                if int(df.loc[index, "Denominator_prev"]) > 50:
                    for sheet_counter in range(len(wb.sheetnames)):
                        if wb.sheetnames[sheet_counter] == 'DenominatorCount':
                            break
                    wb.active = sheet_counter

                    if int(df.loc[index, "Denominator_cur"]) == 0:
                        sh2.append((" ", df.loc[index, "Name_prev"], df.loc[index, "Denominator_prev"], df.loc[index, "Denominator_cur"]))

            elif df.loc[index, '_merge'] == "left_only":
                print(df.loc[index, "Name_prev"] + ": Provider is deleted in the current feed.")
                for sheet_counter in range(len(wb.sheetnames)):
                    if wb.sheetnames[sheet_counter] == 'Deleted':
                        break
                wb.active = sheet_counter
                sh3.append((" ", df.loc[index, "Name_prev"], df.loc[index, "Patients_prev"]))

            elif df.loc[index, '_merge'] == "right_only":
                print(df.loc[index, "Name_cur"] + ": Provider is newly added in the current feed.")
                for sheet_counter in range(len(wb.sheetnames)):
                    if wb.sheetnames[sheet_counter] == 'Added':
                        break
                wb.active = sheet_counter
                sh4.append((" ", df.loc[index, "Name_cur"], df.loc[index, "Patients_cur"]))

        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'PatientsCount':
                break
        wb.active = sheet_counter
        rows = sh1.max_row
        cols = sh1.max_column

        for i in range(2, rows + 1):
            for j in range(5, cols + 1):
                if sh1.cell(i, j).value == 100:
                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
                else:
                    continue
        wb.save(path1 + '\\' + "Report_Provider.xlsx")


c1 = CompareProvidersList(CONFIG_FILE_PATH)
c1.folder_create()
c1.compare_prov_prev_vs_cur()

# if os.path.isfile("C:\\Users\\psur\\Desktop\\Python Selenium Scripts\\Pritam\\ProviderList_HPMG_Prod.csv"):
#    print("File download is completed")
# else:
#    print("File download is not completed")
# df_prev = pd.read_csv("C:\\Users\\psur\\Desktop\\Python Selenium Scripts\\Pritam\\Prospect_CERT.csv", skiprows = 1)
# # print(df_prev)

# df_cur = pd.read_csv("C:\\Users\\psur\\Desktop\\Python Selenium Scripts\\Pritam\\Prospect_Prod.csv", skiprows = 1)
# # print(df_cur)

# df_prev = df_prev.head()
# df_cur = df_cur.head()

# for index in range(len(df_prev)): 
#     flag = 0   
#     for index1 in range(len(df_cur)):
#         if (df_prev.loc[index,"Name"]+str(df_prev.loc[index,"NPI"])) == (df_cur.loc[index1,"Name"]+str(df_cur.loc[index1,"NPI"])):
#             # print(df_prev.loc[index,"Name"]+" PatientCount: "+ str(df_cur.loc[index1,"Patients"]))
#             if int(df_prev.loc[index,"Patients"]) >= 50 and int(df_cur.loc[index1,"Patients"]) == 0:
#                 print(df_prev.loc[index,"Name"]+ ": Patient count becomes zero!")
#             elif int(df_prev.loc[index,"Denominator"]) >= 50 and int(df_cur.loc[index1,"Denominator"]) == 0:
#                 print(df_prev.loc[index,"Name"]+ ": Denominator count becomes zero!")
#             flag = 0
#             break
#         else:
#             flag = 1
#     if flag == 1:
#         print(df_prev.loc[index,"Name"]+" NOT FOUND!")   

# df = pd.merge(df_prev,df_cur,on=' Provider ID',how="outer",indicator=True,suffixes=('_prev','_cur'))
# print(df)

# for index in range(len(df)):    
#     if df.loc[index,'_merge'] == "both":
#         if int(df.loc[index,"Patients_prev"]) >= 50 and int(df.loc[index,"Patients_cur"]) == 0:
#             print(df.loc[index,"Name_prev"]+ ": Patient count becomes zero!")

#         if int(df.loc[index,"Denominator_prev"]) >= 50 and int(df.loc[index,"Denominator_cur"]) == 0:
#             print(df.loc[index,"Name_prev"]+ ": Denominator count becomes zero!")

#         if int(df.loc[index,"Patients_prev"]) > 0 and float((int(df.loc[index,"Patients_prev"]) - int(df.loc[index,"Patients_cur"]))*100/int(df.loc[index,"Patients_prev"])) >= 20:
#             drop_percentage = round(float((int(df.loc[index,"Patients_prev"]) - int(df.loc[index,"Patients_cur"]))*100/int(df.loc[index,"Patients_prev"])))
#             # print(df.loc[index,"Name_prev"] + ": Patient count dropped by "+str(drop_percentage) + "%")
#             if drop_percentage > 20 and drop_percentage < 50:
#                 provider_patient_drop_20to50 = str(df.loc[index,"Name_prev"]) + "|" + str(round(df.loc[index,"Patients_prev"])) + "|" + str(round(df.loc[index,"Patients_cur"])) + "|" +str(drop_percentage)+"%"
#                 print(provider_patient_drop_20to50)

#             elif drop_percentage >= 50 and drop_percentage <= 100:
#                 provider_patient_drop_50to100 = str(df.loc[index,"Name_prev"]) + "|" + str(round(df.loc[index,"Patients_prev"])) + "|" + str(round(df.loc[index,"Patients_cur"])) + "|" +str(drop_percentage)+"%"
#                 print(provider_patient_drop_50to100)

#     elif df.loc[index,'_merge'] == "left_only":
#         print(df.loc[index,"Name_prev"]+ ": Provider is deleted in the current feed.")

#     elif df.loc[index,'_merge'] == "right_only":
#         print(df.loc[index,"Name_cur"]+ ": Provider is newly added in the current feed.")
