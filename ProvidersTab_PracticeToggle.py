import base64
import os
import random
from datetime import date, datetime
import pytz
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from termcolor import colored
from selenium.webdriver.common.action_chains import ActionChains
from configparser import ConfigParser
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
import re

# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
#                                                   USER INPUT
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
CONFIG_FILE_PATH = "G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini"
CUSTOMER_ID = [150, 200, 1000, 1100, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900, 2000, 2100, 2200, 2400, 2500, 2600,
               2700, 2800, 2900, 3000, 3100, 3200, 3300, 3400, 3500, 3600, 3700, 3800, 3900, 4000, 4100, 4200, 4300,
               4400, 4500, 4600, 4700, 4800, 4900, 5000, 5100, 5200, 5300, 5400, 5600, 6000, 6500]


# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


def ProvidersTab():
    """ CONFIG PARSER """
    file = '../config.ini'
    config = ConfigParser()
    config.read(CONFIG_FILE_PATH)

    """ CHROME DRIVER SETUP """
    options = webdriver.ChromeOptions()
    options.add_argument(config['path']['chrome_profile'])  # Path to your chrome profile
    # options.add_argument("--headless")
    # options.add_argument('--disable-gpu')
    # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--start-maximized")
    driver = webdriver.Chrome(executable_path=(config['path']['chrome_driver']), options=options)

    def date_time():
        today = date.today()
        tz_In = pytz.timezone('Asia/Kolkata')
        datetime_In = datetime.now(tz_In)
        time_now = datetime_In.strftime("[%I-%M-%S %p]")
        now = str(today) + time_now
        print(now)
        return now

    wb = Workbook()
    wb.active.title = "ProvidersTab "
    sh1 = wb.active
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    sh1['A1'].font = header_font
    sh1['A1'].fill = header_cell_color
    sh1['B1'].font = header_font
    sh1['B1'].fill = header_cell_color
    sh1['C1'].font = header_font
    sh1['C1'].fill = header_cell_color
    sh1['D1'].font = header_font
    sh1['D1'].fill = header_cell_color
    sh1['E1'].font = header_font
    sh1['E1'].fill = header_cell_color
    sh1['F1'].font = header_font
    sh1['F1'].fill = header_cell_color
    sh1['G1'].font = header_font
    sh1['G1'].fill = header_cell_color

    sh1['A1'] = 'Customer Name'
    sh1['B1'] = 'Tab Name'
    sh1['C1'] = 'Count'
    sh1['D1'] = 'Time Taken'
    sh1['E1'] = 'Status'
    sh1['F1'] = 'Comments'
    sh1['G1'] = 'URL'

    sh1.name = "Arial"

    # FOLDER CREATION
    dateandtime = date_time()
    path = os.path.join((config['path']['parent_dir']), dateandtime + "_ProvidersTab")
    os.mkdir(path)
    path1 = os.path.join(path, "ProvidersTab")
    os.mkdir(path1)
    path2 = os.path.join(path, "Screenshots")
    os.mkdir(path2)

    # LOGGING AREA
    LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
    logging.basicConfig(filename=path1 + "\\" + "Info.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    wb.save(path1 + '\\' + "Report.xlsx")

    def Login():
        try:
            driver.get(config['prod']['logout_url'])
            driver.get(config['prod']['login_url'])
            driver.maximize_window()
            driver.find_element_by_id("edit-name").send_keys(config['credentials']['username'])
            driver.find_element_by_id("edit-pass").send_keys(config['credentials']['password'])
            driver.find_element_by_id("edit-submit").click()
            time.sleep(2)
            WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            driver.find_element_by_id("reason_textbox").send_keys(config['credentials']['reason'])
            time.sleep(0.5)
            driver.find_element_by_id("edit-submit").click()
            ajax_preloader_wait()
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_filterMeasureList'])))
            logger.info("Logged in to Cozeva!")

        except Exception as e:
            print(e)
            logger.critical("Exception occurred in Login function!")
            raise

    def PracticeToggleFromProvidersTab():
        for cust_id_counter in CUSTOMER_ID:
            open_registry_page(cust_id_counter)
            ajax_preloader_wait()
            customer_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
            print("**** " + customer_name + " ****")
            providersTab_url = driver.current_url

            """******************************** PROVIDERS TAB *****************************************"""
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_sidenavSlideOut'])))
            driver.find_element_by_xpath(config['LOCATOR']['xpath_sidenavSlideOut']).click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_providersTab'])))
            time.sleep(0.5)
            driver.find_element_by_xpath(config['LOCATOR']['xpath_providersTab']).click()
            start_time = None
            start_time = time.perf_counter()
            ajax_preloader_wait()
            time_taken_providersTab = round((time.perf_counter() - start_time - 2), 2)
            try:
                provider_count = len(driver.find_elements_by_xpath("//tbody/tr[@role='row']"))
                prov_datatable_info = driver.find_element_by_xpath("//div[@id='metric-support-prov-ls_info']").text

                if provider_count == 0:
                    sh1.append((customer_name, 'Providers', provider_count, time_taken_providersTab, 'FAILED',
                                'Providers Tab is BLANK!', providersTab_url))
                else:
                    sh1.append(
                        (customer_name, 'Providers', prov_datatable_info, time_taken_providersTab, 'PASSED', '', ''))

                """ ******************************** PRACTICE TOGGLE ***************************************** """
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_practiceTab'])))
                driver.find_element_by_xpath(config['LOCATOR']['xpath_practiceTab']).click()
                start_time = None
                start_time = time.perf_counter()
                ajax_preloader_wait()
                time_taken_practiceToggle = round((time.perf_counter() - start_time - 2), 2)

                practiceToggle_url = driver.current_url
                practice_count = len(driver.find_elements_by_xpath("//tbody/tr[@role='row']"))
                prac_datatable_info = driver.find_element_by_xpath("//div[@id='metric-support-prac-ls_info']").text

                if practice_count == 0:
                    sh1.append((customer_name, 'Practices', practice_count, time_taken_practiceToggle, 'FAILED',
                                'Providers Tab is BLANK!', practiceToggle_url))
                else:
                    sh1.append((customer_name, 'Practices', prac_datatable_info, time_taken_practiceToggle, 'PASSED', '', ''))

            except Exception as e:
                print(e)
                sh1.append((customer_name, '', '', '', 'FAILED', 'Exception occurred', providersTab_url))

            finally:
                rows = sh1.max_row
                cols = sh1.max_column
                for i in range(2, rows + 1):
                    for j in range(2, cols + 1):
                        if sh1.cell(i, j).value == 'PASSED':
                            sh1.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                        elif sh1.cell(i, j).value == "FAILED":
                            sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')

                wb.save(path1 + '\\' + "Report.xlsx")

    def ajax_preloader_wait():
        try:
            time.sleep(1)
            WebDriverWait(driver, 120).until(
                EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
            time.sleep(1)

        except Exception as e:
            return e

    def open_registry_page(customer_id):
        customer_list_url = []
        sm_customer_id = str(customer_id)
        sm_customer_id = sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        customer_list_url.append(encoded_string)
        for idx, val in enumerate(customer_list_url):
            # driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))
            driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))

    Login()
    PracticeToggleFromProvidersTab()
    driver.get(config['prod']['logout_url'])
    driver.quit()


ProvidersTab()
