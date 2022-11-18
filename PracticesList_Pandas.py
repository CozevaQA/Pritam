import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
from datetime import date, datetime
import pytz
from configparser import ConfigParser

CONFIG_FILE_PATH = "G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini"

class ConfParser:
    def __init__(self, config_file_path):
        self.config_file_path = config_file_path       
        self.file = '../config.ini'
        self.config = ConfigParser()
        self.config.read(self.config_file_path)
        print("<Execution Log> Config Parser read data from:",self.config_file_path+"\n")

class FolderCreation(ConfParser):   
    def date_time(self):
        self.today = date.today()
        self.tz_In = pytz.timezone('Asia/Kolkata')
        self.datetime_In = datetime.now(self.tz_In)
        self.time = self.datetime_In.strftime("[%I-%M-%S %p]")
        self.now = str(self.today) + self.time
        print("<Execution Log> ",self.now)  
        return self.now            

    def folder_create(self):
        global sh1, sh2, sh3, sh4, wb, path1
        wb = Workbook()
        wb.active.title = "PatientsCount"
        sh1 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh1['A1'].font = header_font
        sh1['A1'].fill = header_cell_color
        sh1['B1'].font = header_font
        sh1['B1'].fill = header_cell_color
        sh1['C1'].font = header_font
        sh1['C1'].fill = header_cell_color
        sh1['D1'].font = header_font
        sh1['D1'].fill = header_cell_color
        sh1['E1'].font = header_font
        sh1['E1'].fill = header_cell_color
        sh1['A1'] = 'Customer Name'
        sh1['B1'] = 'Practice Name'
        sh1['C1'] = 'Previous Patient Count'
        sh1['D1'] = 'Current Patient Count'
        sh1['E1'] = '% Drop'        
        sh1.name = "Arial"

        sh2 = wb.create_sheet(title='DenominatorCount')
        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'DenominatorCount':
                break
        wb.active = sheet_counter
        sh2 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh2['A1'].font = header_font
        sh2['A1'].fill = header_cell_color
        sh2['B1'].font = header_font
        sh2['B1'].fill = header_cell_color
        sh2['C1'].font = header_font
        sh2['C1'].fill = header_cell_color
        sh2['D1'].font = header_font
        sh2['D1'].fill = header_cell_color
        sh2['A1'] = 'Customer Name'
        sh2['B1'] = 'Practice Name'
        sh2['C1'] = 'Previous Denominator'
        sh2['D1'] = 'Current Denominator'      
        sh2.name="Arial"

        sh3 = wb.create_sheet(title='Deleted')
        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'Deleted':
                break
        wb.active = sheet_counter
        sh3 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh3['A1'].font = header_font
        sh3['A1'].fill = header_cell_color
        sh3['B1'].font = header_font
        sh3['B1'].fill = header_cell_color
        sh3['C1'].font = header_font
        sh3['C1'].fill = header_cell_color        
        sh3['A1'] = 'Customer Name'
        sh3['B1'] = 'Practice Name'
        sh3['C1'] = 'Previous Patients Count'           
        sh3.name="Arial"

        sh4 = wb.create_sheet(title='Added')
        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'Added':
                break
        wb.active = sheet_counter
        sh4 = wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        sh4['A1'].font = header_font
        sh4['A1'].fill = header_cell_color
        sh4['B1'].font = header_font
        sh4['B1'].fill = header_cell_color
        sh4['C1'].font = header_font
        sh4['C1'].fill = header_cell_color        
        sh4['A1'] = 'Customer Name'
        sh4['B1'] = 'Practice Name'
        sh4['C1'] = 'Current Patients Count'           
        sh4.name="Arial"

        self.dateandtime = self.date_time()
        print("D1:",self.config_file_path)        
        path = os.path.join((self.config['path']['pandas_dir']), self.dateandtime)
        os.mkdir(path)
        path1 = os.path.join(path, " Report_Practice")    
        os.mkdir(path1)
        wb.save(path1 + '\\' + " Report_Practice.xlsx")
        print("<Execution Log> Folder Created")

class ComparePracticeList(FolderCreation):
    def compare_prac_prev_vs_cur(self):
        df_prev = pd.read_csv("C:\\Users\\psur\\Documents\\PandasPerformance\\PandasInput\\Practice\\PracticeList_StJoseph_16Nov.csv", skiprows = 1)
        df_prev = df_prev.drop_duplicates(subset='Name', keep="first")
        print(df_prev)

        df_cur = pd.read_csv("C:\\Users\\psur\\Documents\\PandasPerformance\\PandasInput\\Practice\\PracticeList_StJoseph_17Nov.csv", skiprows = 1)
        df_cur = df_cur.drop_duplicates(subset='Name', keep="first")
        print(df_cur)

        df = pd.merge(df_prev,df_cur,on='Name',how="outer",indicator=True,suffixes=('_prev','_cur'))
        print(df)

        for index in range(len(df)):    
            if df.loc[index,'_merge'] == "both":        

                if int(df.loc[index,"Patients_prev"]) > 50:
                    for sheet_counter in range(len(wb.sheetnames)):
                        if wb.sheetnames[sheet_counter] == 'PatientsCount':
                            break
                    wb.active = sheet_counter
                    drop_percentage = round(float((int(df.loc[index,"Patients_prev"]) - int(df.loc[index,"Patients_cur"]))*100/int(df.loc[index,"Patients_prev"])))   

                    if int(df.loc[index,"Patients_cur"]) == 0:
                        print(df.loc[index,"Name"]+ ": Patient count becomes zero!")                        
                        sh1.append((" ", df.loc[index,"Name"], df.loc[index,"Patients_prev"], df.loc[index,"Patients_cur"], drop_percentage ))

                    elif drop_percentage > 50 and drop_percentage < 100:
                        sh1.append((" ", df.loc[index,"Name"], df.loc[index,"Patients_prev"], df.loc[index,"Patients_cur"], drop_percentage ))

                if int(df.loc[index,"Denominator_prev"]) > 50:
                    for sheet_counter in range(len(wb.sheetnames)):
                        if wb.sheetnames[sheet_counter] == 'DenominatorCount':
                            break
                    wb.active = sheet_counter

                    if int(df.loc[index,"Denominator_cur"]) == 0:
                        sh2.append((" ", df.loc[index,"Name"], df.loc[index,"Denominator_prev"], df.loc[index,"Denominator_cur"] ))                    

            elif df.loc[index,'_merge'] == "left_only":
                print(df.loc[index,"Name"]+ ": Practice is deleted in the current feed.")
                for sheet_counter in range(len(wb.sheetnames)):
                        if wb.sheetnames[sheet_counter] == 'Deleted':
                            break
                wb.active = sheet_counter
                sh3.append((" ", df.loc[index,"Name"], df.loc[index,"Patients_prev"]))

            elif df.loc[index,'_merge'] == "right_only":
                print(df.loc[index,"Name"]+ ": Practice is newly added in the current feed.")
                for sheet_counter in range(len(wb.sheetnames)):
                        if wb.sheetnames[sheet_counter] == 'Added':
                            break
                wb.active = sheet_counter
                sh4.append((" ", df.loc[index,"Name"], df.loc[index,"Patients_cur"]))

        for sheet_counter in range(len(wb.sheetnames)):
            if wb.sheetnames[sheet_counter] == 'PatientsCount':
                break
        wb.active = sheet_counter
        rows = sh1.max_row
        cols = sh1.max_column

        for i in range(2, rows + 1):
            for j in range(5, cols + 1):
                if sh1.cell(i, j).value == 100:
                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')
                else:
                    continue
        wb.save(path1 + '\\' + " Report_Practice.xlsx")


c1 = ComparePracticeList(CONFIG_FILE_PATH)
c1.folder_create()
c1.compare_prac_prev_vs_cur()