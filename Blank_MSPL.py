import base64
import os
import random
from datetime import date, datetime
import pytz
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from termcolor import colored
from selenium.webdriver.common.action_chains import ActionChains
from configparser import ConfigParser
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
import re


# """ USER INPUT """
# CONFIG_FILE_PATH = "C:\\Users\\psur\\PycharmProjects\\Cozeva_Automation\\config.ini"
# CUSTOMER_ID = 2100

def Blank_MSPL_Script(CUSTOMER_ID, QUARTER, COUNTER):
    """"
    ********************************************************* BLANK MSPL (Provider's context) VALIDATION SCRIPT *********************************************************
    """

    """ CONFIG PARSER """
    file = '../config.ini'
    config = ConfigParser()
    config.read("G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini")

    """ CHROME DRIVER SETUP """
    options = webdriver.ChromeOptions()
    options.add_argument(config['path']['chrome_profile'])  # Path to your chrome profile
    # options.add_argument("--headless")
    # options.add_argument('--disable-gpu')
    # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--start-maximized")
    driver = webdriver.Chrome(executable_path=(config['path']['chrome_driver']), options=options)

    # **** CURRENT DATE & TIME ****
    def date_time():
        today = date.today()
        tz_In = pytz.timezone('Asia/Kolkata')
        datetime_In = datetime.now(tz_In)
        time_now = datetime_In.strftime("[%I-%M-%S %p]")
        now = str(today) + time_now
        print(now)
        return now

    wb = Workbook()
    wb.active.title = "CUSTOMER_ID " + str(CUSTOMER_ID)
    sh1 = wb.active
    header_font = Font(color='FFFFFF', bold=False, size=12)
    header_cell_color = PatternFill('solid', fgColor='030303')
    sh1['A1'].font = header_font
    sh1['A1'].fill = header_cell_color
    sh1['B1'].font = header_font
    sh1['B1'].fill = header_cell_color
    sh1['C1'].font = header_font
    sh1['C1'].fill = header_cell_color
    sh1['D1'].font = header_font
    sh1['D1'].fill = header_cell_color
    sh1['E1'].font = header_font
    sh1['E1'].fill = header_cell_color
    sh1['F1'].font = header_font
    sh1['F1'].fill = header_cell_color
    sh1['A1'] = 'Quarter & LOB'
    sh1['B1'] = 'Measure Name'
    sh1['C1'] = 'Provider Name'
    sh1['D1'] = 'Patient Count'
    sh1['E1'] = 'Status'
    sh1['F1'] = 'URL'
    sh1.name = "Arial"

    # **** FOLDER CREATION ****
    dateandtime = date_time()
    path = os.path.join((config['path']['parent_dir']), dateandtime + "_Provider MSPL_Customer " + str(CUSTOMER_ID))
    os.mkdir(path)
    path1 = os.path.join(path, "Cust Id " + str(CUSTOMER_ID) + "_" + "MSPL")
    os.mkdir(path1)
    path2 = os.path.join(path, "Screenshots")
    os.mkdir(path2)

    # **** LOGGING AREA ****
    LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
    logging.basicConfig(filename=path1 + "\\" + "Info.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    wb.save(path1 + '\\' + "Report.xlsx")

    def Login():
        try:
            driver.get(config['prod']['logout_url'])
            driver.get(config['prod']['login_url'])
            driver.maximize_window()
            driver.find_element_by_id("edit-name").send_keys(config['credentials']['username'])
            driver.find_element_by_id("edit-pass").send_keys(config['credentials']['password'])
            driver.find_element_by_id("edit-submit").click()
            time.sleep(2)
            WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            driver.find_element_by_id("reason_textbox").send_keys(config['credentials']['reason'])
            time.sleep(0.5)
            driver.find_element_by_id("edit-submit").click()
            ajax_preloader_wait()
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, config['LOCATOR']['xpath_filterMeasureList'])))
            logger.info("Logged in to Cozeva!")

        except Exception as e:
            print(e)
            logger.critical("Exception occurred in Login function!")
            raise

    def mspl_blank():
        open_registry_page(CUSTOMER_ID)
        time.sleep(5)
        ajax_preloader_wait()
        customer_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
        print("**** " + customer_name + " ****")
        driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
        time.sleep(1)
        quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")
        lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
        driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()

        """    
        *** QUARTER & LOB SELECTION ***    
        """
        for quarter in range(QUARTER):
            # quarter = quarter + 1
            for lob in range(len(lobs)):
                # for lob in range(3):
                # lob = lob + 3
                time.sleep(0.5)
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@id='qt-filter-label']")))
                driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
                time.sleep(0.25)
                quarter_name = quarters[quarter].text
                print(colored(quarter_name, 'blue'))
                quarters[quarter].click()
                time.sleep(0.25)
                lobs[lob].click()
                lob_name = lobs[lob].text
                lob_name = re.sub("[\/:*?<>|]", "", lob_name)
                print(colored(lob_name, 'magenta'))
                driver.find_element_by_xpath("//a[@id='reg-filter-apply']").click()
                ajax_preloader_wait()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))
                driver.find_element_by_xpath("//a[@data-target='qt-reg-nav-filters']").click()
                time.sleep(0.25)
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//label[@class='col s12']")))
                driver.find_element_by_xpath("//label[@class='col s12']").click()
                time.sleep(0.25)
                driver.find_element_by_xpath("//button[@id='qt-apply-search']").click()
                ajax_preloader_wait()

                """
                **** SUPPORT MEASURE REGISTRY NAVIGATION ****
                """
                measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                measure_counter = 0
                score = 0

                while measure_counter < len(measures_all) and score < len(scores):
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.XPATH, "//a[@id='reg-faq-trigger']")))
                    time.sleep(0.5)

                    driver.execute_script("arguments[0].scrollIntoView();", measures_all[measure_counter])
                    measure_name = (measures_all[measure_counter]).text
                    print("Measure name: ", measure_name)
                    measure_name = re.sub("[\/:*?<>|]", "", measure_name)
                    numdeno = scores[score].text
                    numdeno = numdeno.lstrip("(")
                    numdeno = numdeno.rstrip(")")
                    numdeno = numdeno.split("/")
                    Numerator = numdeno[0].replace(',', '')
                    Denominator = numdeno[1].replace(',', '')
                    print("Numerator=", Numerator)
                    print("Denominator=", Denominator)
                    last_url = driver.current_url
                    measures_all[measure_counter].click()
                    try:
                        ajax_preloader_wait()
                        if len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0 and float(
                                Denominator) != 0 and float(Numerator) != 0:
                            met_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", measure_name)
                            logger.warning("Providers list is blank. Please check manually.")
                            sh1.append(
                                (quarter_name + " | " + lob_name, measure_name, 'Blank Providers List', 'NA', 'FAILED'))
                            driver.save_screenshot(
                                path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + ".png")

                        elif len(
                                driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0 and float(
                            Denominator) == 0:
                            met_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", measure_name)
                            logger.info("Providers list is blank since measure score is zero.")
                            sh1.append((quarter_name + " | " + lob_name, measure_name,
                                        'Providers list is blank since measure score is zero.', 'NA', 'NA'))
                            driver.save_screenshot(
                                path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + ".png")

                        elif len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0:
                            met_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", measure_name)
                            logger.info("Providers list is blank. Please check manually.")
                            sh1.append((quarter_name + " | " + lob_name, measure_name,
                                        'Providers list is blank. Please check manually.', 'NA', 'NA'))
                            driver.save_screenshot(
                                path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + ".png")

                        else:
                            if len(driver.find_elements_by_xpath("(//td/a[contains(@href,'/registries/')])[2]")) != 0:
                                patient_list_link = driver.find_element_by_xpath(
                                    "(//td/a[contains(@href,'/registries/')])[2]")
                                provider_name = driver.find_element_by_xpath(
                                    "(//td/a[contains(@href,'/registries?')])[2]").text
                            else:
                                patient_list_link = driver.find_element_by_xpath(
                                    "(//td/a[contains(@href,'/registries/')])[1]")
                                provider_name = driver.find_element_by_xpath(
                                    "(//td/a[contains(@href,'/registries?')])[1]").text

                            ActionChains(driver).move_to_element(patient_list_link).perform()
                            ActionChains(driver).key_down(Keys.CONTROL).click(patient_list_link).key_up(
                                Keys.CONTROL).perform()
                            driver.switch_to.window(driver.window_handles[1])
                            time.sleep(0.5)
                            mspl_url0 = driver.current_url

                            try:
                                ajax_preloader_wait()
                                driver.find_element_by_xpath(
                                    "//a[@class='datatable_filter_dropdown sidenav-trigger']").click()
                                time.sleep(1)
                                WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.LINK_TEXT, 'Apply')))
                                driver.find_element_by_link_text('Clear All').click()
                                time.sleep(0.5)
                                driver.find_element_by_link_text('Apply').click()
                                ajax_preloader_wait()
                                patient_count = len(
                                    driver.find_elements_by_xpath("//td/div/a[contains(@href,'/patient_detail/')]"))
                                mspl_url1 = driver.current_url

                                if patient_count == 0:
                                    sh1.append((quarter_name + " | " + lob_name, measure_name, provider_name,
                                                'MSPL is blank!!', 'FAILED', mspl_url1))
                                elif patient_count != 0:
                                    datatable_info = driver.find_element_by_xpath(
                                        "//div[@id='quality_registry_list_info']").text
                                    sh1.append((quarter_name + " | " + lob_name, measure_name, provider_name,
                                                datatable_info, 'PASSED'))
                                    driver.save_screenshot(
                                        path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + ".png")

                            except Exception as e:
                                print(e)
                                logger.critical(
                                    measure_name + '\n' + provider_name + '\n' + "Metric specific patients list is not opening!Exception occurred!!")
                                sh1.append((quarter_name + " | " + lob_name, measure_name, provider_name,
                                            'MSPL is NOT opening!', 'FAILED', mspl_url0))
                                driver.save_screenshot(
                                    path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + ".png")

                            finally:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])

                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//a[@class='breadcrumb']")))
                        driver.find_element_by_xpath("//a[@class='breadcrumb']").click()

                    # Providers list open exception block
                    except Exception as e:
                        print(e)
                        sh1.append((quarter_name + " | " + lob_name, measure_name, 'Providers list is not opening!',
                                    "NA", 'FAILED'))
                        driver.save_screenshot(
                            path2 + "\\" + quarter_name + " " + lob_name + "_" + measure_name + ".png")
                        driver.get(last_url)
                        ajax_preloader_wait()
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))
                        driver.find_element_by_xpath("//a[@data-target='qt-reg-nav-filters']").click()
                        time.sleep(0.25)
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//label[@class='col s12']")))
                        driver.find_element_by_xpath("//label[@class='col s12']").click()
                        time.sleep(0.25)
                        driver.find_element_by_xpath("//button[@id='qt-apply-search']").click()
                        ajax_preloader_wait()

                    finally:
                        rows = sh1.max_row
                        cols = sh1.max_column
                        for i in range(2, rows + 1):
                            for j in range(2, cols + 1):
                                if sh1.cell(i, j).value == 'PASSED':
                                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                                elif sh1.cell(i, j).value == "FAILED":
                                    sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')

                        wb.save(path1 + '\\' + "Report.xlsx")

                        # MEASURE COUNTER
                        measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                        scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                        measure_counter += COUNTER
                        score += COUNTER

                lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
                quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")

    def open_registry_page(customer_id):
        customer_list_url = []
        sm_customer_id = str(customer_id)
        sm_customer_id = sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
            sm_customer_id) + '&orgId=' + str(sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        customer_list_url.append(encoded_string)
        for idx, val in enumerate(customer_list_url):
            driver.get("https://www.cozeva.com/registries?session=" + val.decode('utf-8'))

    def ajax_preloader_wait():
        try:
            time.sleep(1)
            WebDriverWait(driver, 120).until(
                EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
            time.sleep(1)

        except Exception as e:
            return e

    Login()
    mspl_blank()
    driver.get(config['prod']['logout_url'])
    driver.quit()


Blank_MSPL_Script(200, 3, 10)
Blank_MSPL_Script(2100, 2, 3)
Blank_MSPL_Script(2000, 2, 5)
