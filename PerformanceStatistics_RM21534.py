import base64
import os
import random
from datetime import date, datetime
from unittest import skip
import pytz
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from termcolor import colored
from selenium.webdriver.common.action_chains import ActionChains
from configparser import ConfigParser
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
import re

CONFIG_FILE_PATH = "G:\\My Drive\\IMPORTANT\\PYTHON&PYCHARM\\PycharmProjects\\Cozeva_Automation\\config.ini"

class ConfParser:
    def __init__(self, config_file_path):
        self.config_file_path = config_file_path       
        self.file = '../config.ini'
        self.config = ConfigParser()
        self.config.read(self.config_file_path)
        print("<Execution Log> Config Parser read data from:",self.config_file_path+"\n")

class ChromeDriverSetup(ConfParser):
    def __init__(self, config_file_path):
        ConfParser.__init__(self, config_file_path)
        self.options = webdriver.ChromeOptions()
        self.options.add_argument(self.config['path']['chrome_profile'])  # Path to your chrome profile
        # options.add_argument("--headless")
        # options.add_argument('--disable-gpu')
        # options.add_argument("--window-size=1920,1080")
        # options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(executable_path=(self.config['path']['chrome_driver']), options=self.options)
        print("<Execution Log> Chrome Driver Setup done!\n")

class FolderCreation(ConfParser):   
    def date_time(self):
        self.today = date.today()
        self.tz_In = pytz.timezone('Asia/Kolkata')
        self.datetime_In = datetime.now(self.tz_In)
        self.time = self.datetime_In.strftime("[%I-%M-%S %p]")
        self.now = str(self.today) + self.time
        print("<Execution Log> ",self.now)  
        return self.now            

    def folder_create(self):
        self.wb = Workbook()
        self.wb.active.title = "PerformanceStatistics"
        self.sh1 = self.wb.active
        header_font = Font(color='FFFFFF', bold=False, size=12)
        header_cell_color = PatternFill('solid', fgColor='030303')
        self.sh1['A1'].font = header_font
        self.sh1['A1'].fill = header_cell_color
        self.sh1['B1'].font = header_font
        self.sh1['B1'].fill = header_cell_color
        self.sh1['C1'].font = header_font
        self.sh1['C1'].fill = header_cell_color
        self.sh1['D1'].font = header_font
        self.sh1['D1'].fill = header_cell_color
        self.sh1['E1'].font = header_font
        self.sh1['E1'].fill = header_cell_color
        self.sh1['F1'].font = header_font
        self.sh1['F1'].fill = header_cell_color
        self.sh1['A1'] = 'Customer Name'
        self.sh1['B1'] = 'Quarter & LOB'
        self.sh1['C1'] = 'Measure Name'
        self.sh1['D1'] = 'Previous value in Default quarter'
        self.sh1['E1'] = 'Current value in Previous quarter'
        self.sh1['F1'] = 'Status'
        self.sh1.name = "Arial"


        self.dateandtime = self.date_time()
        print("D1:",self.config_file_path)        
        self.path = os.path.join((self.config['path']['parent_dir']), self.dateandtime)
        os.mkdir(self.path)
        self.path1 = os.path.join(self.path, "Report")    
        os.mkdir(self.path1)
        self.wb.save(self.path1 + '\\' + "Report.xlsx")
        print("<Execution Log> Folder Created\n")

class SupportiveFunctions:
    def ajax_preloader_wait(self):
        try:
            time.sleep(1)
            WebDriverWait(self.driver, 300).until(EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
            time.sleep(1)
        except Exception as e:
            return e   

    
       
class CozevaLogin(ChromeDriverSetup, FolderCreation, SupportiveFunctions):       
    def login_cozeva(self):
        try:
            self.driver.get(self.config['cert']['logout_url'])
            self.driver.get(self.config['cert']['login_url'])
            self.driver.maximize_window()
            self.driver.find_element_by_id("edit-name").send_keys(self.config['credentials']['username'])
            self.driver.find_element_by_id("edit-pass").send_keys(self.config['credentials']['password'])
            self.driver.find_element_by_id("edit-submit").click()
            time.sleep(2)
            WebDriverWait(self.driver, 90).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
            self.driver.find_element_by_id("reason_textbox").send_keys(self.config['credentials']['reason'])
            time.sleep(0.5)
            self.driver.find_element_by_id("edit-submit").click()
            self.ajax_preloader_wait()
            WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, self.config['LOCATOR']['xpath_filterMeasureList'])))
            print("<Execution Log> Logged in to COZEVA.\n")            

        except Exception as e:
            print(e)            
            raise

class CustomerRegistryPage(CozevaLogin):
    def open_registry_page(self, customer_id):
        self.customer_list_url = []
        self.sm_customer_id = str(customer_id)
        self.sm_customer_id = self.sm_customer_id.split(".")[0]
        session_var = 'app_id=registries&custId=' + str(self.sm_customer_id) + '&payerId=' + str(self.sm_customer_id) + '&orgId=' + str(self.sm_customer_id)
        encoded_string = base64.b64encode(session_var.encode('utf-8'))
        self.customer_list_url.append(encoded_string)
        for self.idx, self.val in enumerate(self.customer_list_url):
            self.driver.get("https://cert.cozeva.com/registries?session=" + self.val.decode('utf-8'))
        self.ajax_preloader_wait()
        time.sleep(5)
        self.customer_name = self.driver.find_element_by_xpath(self.config['LOCATOR']['xpath_contextName']).text
        self.registry_url = self.driver.current_url
        print(self.registry_url)
        print("<Execution Log> Support Registry page navigated for ",self.customer_name)  

class SupportLevel(CustomerRegistryPage):
    
    def quarter_lob_traverse(self):
        WebDriverWait(self.driver, 90).until(EC.presence_of_element_located((By.XPATH, "//a[@id='qt-filter-label']")))
        self.driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
        time.sleep(2)
        WebDriverWait(self.driver, 90).until(EC.presence_of_element_located((By.XPATH, "//ul[@id='filter-quarter']/li")))
        quarters = self.driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")
        lobs = self.driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
        self.driver.find_element_by_xpath("//a[@id='qt-filter-label']").click() 
        # CE Toggle
        # self.continuous_enrollment_click()

        self.provider_list =None
        self.provider_list =[]
             
        for quarter in range(1):            
            # quarter = quarter + 1            
            for lob in range(len(lobs)):
                # lob = lob + 1               
                print("<DEBUG>Quarter",quarter)
                print("<DEBUG>LoB",lob)
                time.sleep(0.5)
                WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@id='qt-filter-label']")))
                self.driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()  
                time.sleep(2) 
                WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@id='reg-filter-apply']")))                             
                quarter_name = quarters[quarter].text
                print(colored(quarter_name, 'blue'))
                quarters[quarter].click()
                time.sleep(0.25)
                lobs[lob].click()
                lob_name = lobs[lob].text
                lob_name=re.sub("[\/:*?<>|]","",lob_name)
                print(colored(lob_name, 'magenta'))                
                self.driver.find_element_by_xpath("//a[@id='reg-filter-apply']").click()
                self.ajax_preloader_wait()
                WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))


                self.urls_metric_specific_providers_list()
               
            
                self.ajax_preloader_wait()
                lobs = self.driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
                quarters = self.driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")


        print(self.provider_list)

    def continuous_enrollment_click(self):
        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//label[@for='conti_enroll']")))
        self.driver.find_element_by_xpath("//label[@for='conti_enroll']").click()
        time.sleep(1)



                

    
    def urls_metric_specific_providers_list(self):
        time.sleep(5)
        #self.providers_list_links = []
        self.providers_list_links = self.driver.find_elements_by_xpath("//div[@class='qt-metric']/a")        
        # print(self.providers_list_links)
         
        for self.measure_counter in range(len(self.providers_list_links)):                       
            self.provider_list.append(self.providers_list_links[self.measure_counter].get_attribute("href"))

            self.providers_list_links = self.driver.find_elements_by_xpath("//div[@class='qt-metric']/a")
        # print(self.provider_list)

    def navigation_metric_specific_providers_list(self):
        for self.provider_list_counter in self.provider_list:
            try:
                print("<Execution Log> ", self.provider_list_counter)
                self.driver.get(self.provider_list_counter)
                self.ajax_preloader_wait()
            except Exception as e:
                print(e)

        self.driver.get(self.registry_url)


    def navigation_metric_specific_patients_list(self):
        for self.provider_list_counter in self.provider_list:
            try:
                self.patient_list_counter = self.provider_list_counter + "&table_id=metric-support-pat-ls"
                print("<Execution Log> ", self.patient_list_counter)
                self.driver.get(self.patient_list_counter)
                self.ajax_preloader_wait()
            except Exception as e:
                print(e)
        self.driver.get(self.registry_url)

    def navigation_performance_statistics(self):
        for self.provider_list_counter in self.provider_list:
            try:
                self.performance_statistics_counter = self.provider_list_counter + "&tab_type=registry_performance#registry_performance"
                print("<Execution Log> ", self.performance_statistics_counter)
                self.driver.get(self.performance_statistics_counter)
                
                self.ajax_preloader_wait()
                self.my_function_RM21534()
                
            except Exception as e:
                print(e)
        self.driver.get(self.registry_url)

    def my_function_RM21534(self):
        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//div[@class='ch metric_specific_patient_list_title']")))
        measure_name_default_quarter = self.driver.find_element_by_xpath("//div[@class='ch metric_specific_patient_list_title']").text
        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//span[@class='prev_performance']")))
        previous_value_in_default_quarter = self.driver.find_element_by_xpath("//span[@class='prev_performance']").text
        quarter_lob = self.driver.find_element_by_xpath("//div[@class='metric_patient_list_filter left']").text
        previous_value_in_default_quarter = re.sub("%",'',previous_value_in_default_quarter)
        previous_value_in_default_quarter = float(previous_value_in_default_quarter)

        performance_statistics_counter_previous_quarter = re.sub("2022-12-31", "2021-12-31", self.performance_statistics_counter)
        print(performance_statistics_counter_previous_quarter)        
        self.driver.get(performance_statistics_counter_previous_quarter)
        self.ajax_preloader_wait()

        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//div[@class='ch metric_specific_patient_list_title']")))
        measure_name_previous_quarter = self.driver.find_element_by_xpath("//div[@class='ch metric_specific_patient_list_title']").text
        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//div[@class='performance_value']")))
        current_value_in_previous_quarter = self.driver.find_element_by_xpath("//div[@class='performance_value']").text
        current_value_in_previous_quarter = re.sub("%",'',current_value_in_previous_quarter)
        current_value_in_previous_quarter = float(current_value_in_previous_quarter)

        if measure_name_default_quarter == measure_name_previous_quarter:
            if abs(((previous_value_in_default_quarter - current_value_in_previous_quarter)*100/previous_value_in_default_quarter)) <= 1:
                print(measure_name_default_quarter)
                print(measure_name_previous_quarter)
                print(previous_value_in_default_quarter)
                print(current_value_in_previous_quarter)
                print("MATCH FOUND")
                self.sh1.append((self.customer_name, quarter_lob, measure_name_default_quarter, previous_value_in_default_quarter, current_value_in_previous_quarter, 'PASSED'))
            else:
                print(measure_name_default_quarter)
                print(measure_name_previous_quarter)
                print(previous_value_in_default_quarter)
                print(current_value_in_previous_quarter)
                print("MATCH DOES NOT FOUND")
                self.sh1.append((self.customer_name, quarter_lob, measure_name_default_quarter, previous_value_in_default_quarter, current_value_in_previous_quarter, 'FAILED'))

        elif measure_name_default_quarter != measure_name_previous_quarter:
            print(measure_name_default_quarter)
            print(measure_name_previous_quarter)
            print(previous_value_in_default_quarter)
            print(current_value_in_previous_quarter)
            print("MEASURE PRESENCE DISCREPANCY FOUND")
            self.sh1.append((self.customer_name, quarter_lob, measure_name_default_quarter, previous_value_in_default_quarter,'', 'MEASURE PRESENCE DISCREPANCY'))

        rows = self.sh1.max_row
        cols = self.sh1.max_column

        for i in range(2, rows + 1):
            for j in range(2, cols + 1):
                if self.sh1.cell(i, j).value == 'PASSED':
                    self.sh1.cell(i, j).fill = PatternFill('solid', fgColor='0FC404')
                elif self.sh1.cell(i, j).value == "FAILED":
                    self.sh1.cell(i, j).fill = PatternFill('solid', fgColor='FC0E03')                             
                                
                                
                                

        self.wb.save(self.path1 + '\\' + "Report.xlsx")



        
    
 

            
            





c1 = SupportLevel(CONFIG_FILE_PATH)
c1.folder_create()
c1.login_cozeva()
c1.open_registry_page(4000)
c1.quarter_lob_traverse()
c1.navigation_performance_statistics()